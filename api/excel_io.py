"""
Excel file I/O operations for FastAPI backend.
Reuses logic from the Streamlit version.
"""

import pandas as pd
from pathlib import Path
from typing import Dict, Optional, BinaryIO, Any
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dataclasses import dataclass


@dataclass
class ColumnDefinition:
    """Definition for a single column in the Excel sheet."""
    name: str
    display_name: str
    data_type: str
    required: bool
    max_length: Optional[int] = None
    default_value: Any = None
    unique: bool = False
    help_text: str = ""
    pattern: Optional[str] = None
    pattern_message: str = ""


# ============================================================================
# MODULES SHEET CONFIGURATION
# ============================================================================

MODULES_COLUMNS = {
    "moduleCode": ColumnDefinition(
        name="moduleCode", display_name="Module Code", data_type="string",
        required=True, max_length=64, unique=True,
        help_text="Unique identifier for the module",
        pattern=r"^[A-Z0-9][A-Z0-9\-]*[A-Z0-9]$|^[A-Z0-9]$",
        pattern_message="Module code should be UPPERCASE with hyphens"
    ),
    "isUserAssignable": ColumnDefinition(
        name="isUserAssignable", display_name="User Assignable", data_type="boolean",
        required=True, default_value=True,
        help_text="If TRUE, providers can manually assign this module"
    ),
    "name": ColumnDefinition(
        name="name", display_name="Name", data_type="string",
        required=True, max_length=255,
        help_text="Internal display name for the module"
    ),
    "clientFriendlyName": ColumnDefinition(
        name="clientFriendlyName", display_name="Client Friendly Name", data_type="string",
        required=False,
        help_text="User-friendly name shown to clients"
    ),
    "description": ColumnDefinition(
        name="description", display_name="Description", data_type="string",
        required=False,
        help_text="Detailed description of the module"
    ),
}

MODULES_COLUMN_ORDER = ["moduleCode", "isUserAssignable", "name", "clientFriendlyName", "description"]

# ============================================================================
# ASSESSMENTS SHEET CONFIGURATION
# ============================================================================

ASSESSMENTS_COLUMNS = {
    "level": ColumnDefinition(
        name="level", display_name="Level", data_type="integer",
        required=True, default_value=1,
        help_text="Assessment level (0-3)"
    ),
    "assessmentCode": ColumnDefinition(
        name="assessmentCode", display_name="Assessment Code", data_type="string",
        required=True, max_length=64, unique=True,
        help_text="Unique identifier",
        pattern=r"^[A-Z0-9][A-Z0-9_]*[A-Z0-9]$|^[A-Z0-9]+$",
        pattern_message="Assessment code should be UPPERCASE with underscores"
    ),
    "moduleCode": ColumnDefinition(
        name="moduleCode", display_name="Module Code", data_type="string",
        required=False, max_length=64,
        help_text="Links to module that executes this assessment"
    ),
    "isUserAssignable": ColumnDefinition(
        name="isUserAssignable", display_name="User Assignable", data_type="boolean",
        required=True, default_value=False,
        help_text="If TRUE, providers can manually assign"
    ),
    "name": ColumnDefinition(
        name="name", display_name="Name", data_type="string",
        required=True, max_length=255,
        help_text="Internal display name"
    ),
    "clientFriendlyName": ColumnDefinition(
        name="clientFriendlyName", display_name="Client Friendly Name", data_type="string",
        required=False,
        help_text="Client-facing name"
    ),
    "description": ColumnDefinition(
        name="description", display_name="Description", data_type="string",
        required=False,
        help_text="Detailed description"
    ),
    "estimatedDuration": ColumnDefinition(
        name="estimatedDuration", display_name="Estimated Duration", data_type="string",
        required=False, max_length=64,
        help_text="Expected completion time"
    ),
}

ASSESSMENTS_COLUMN_ORDER = [
    "level", "assessmentCode", "moduleCode", "isUserAssignable",
    "name", "clientFriendlyName", "description", "estimatedDuration"
]

# ============================================================================
# LEVEL FACTS SHEET CONFIGURATION
# ============================================================================

LEVEL_FACTS_COLUMNS = {
    "assessmentId": ColumnDefinition(
        name="assessmentId", display_name="Assessment ID", data_type="string",
        required=True,
        help_text="Links fact to its parent assessment"
    ),
    "nodeId": ColumnDefinition(
        name="nodeId", display_name="Node ID", data_type="string",
        required=True,
        help_text="Unique identifier for this question node"
    ),
    "factId": ColumnDefinition(
        name="factId", display_name="Fact ID", data_type="string",
        required=False,
        help_text="Unique identifier for the collected data point"
    ),
    "factGroup": ColumnDefinition(
        name="factGroup", display_name="Fact Group", data_type="string",
        required=False,
        help_text="Groups related facts for algorithm processing"
    ),
    "nodeText": ColumnDefinition(
        name="nodeText", display_name="Node Text", data_type="string",
        required=True,
        help_text="The question or statement text"
    ),
    "dataType": ColumnDefinition(
        name="dataType", display_name="Data Type", data_type="enum",
        required=True, default_value="enumeration",
        help_text="Type of response"
    ),
    "isDeterministic": ColumnDefinition(
        name="isDeterministic", display_name="Is Deterministic", data_type="boolean",
        required=True, default_value=True,
        help_text="TRUE = uses deterministic code/buttons"
    ),
    "range": ColumnDefinition(
        name="range", display_name="Range", data_type="string",
        required=False,
        help_text="Valid value range for integer dataType",
        pattern=r"^\d+-\d+$",
        pattern_message="Range must be in format 'min-max'"
    ),
    "enumerationType": ColumnDefinition(
        name="enumerationType", display_name="Enumeration Type", data_type="string",
        required=False,
        help_text="References enumeration definition"
    ),
    "target": ColumnDefinition(
        name="target", display_name="Target", data_type="string",
        required=False,
        help_text="Next node(s) to route to"
    ),
    "isRestartPoint": ColumnDefinition(
        name="isRestartPoint", display_name="Is Restart Point", data_type="boolean",
        required=True, default_value=False,
        help_text="TRUE = session can resume from this point"
    ),
}

LEVEL_FACTS_COLUMN_ORDER = [
    "assessmentId", "nodeId", "factId", "factGroup", "nodeText", "dataType",
    "isDeterministic", "range", "enumerationType", "target", "isRestartPoint"
]

LEVEL_FACTS_INFO = {
    0: {"name": "Level 0 Facts", "description": "Introduction and onboarding questions", "icon": "🏠"},
    1: {"name": "Level 1 Facts", "description": "Universal screening questions", "icon": "🔍"},
    2: {"name": "Level 2 Facts", "description": "Focused domain assessment questions", "icon": "🎯"},
    3: {"name": "Level 3 Facts", "description": "Detailed diagnostic questions", "icon": "🔬"},
}

# ============================================================================
# ALGORITHMS SHEET CONFIGURATION
# ============================================================================

ALGORITHMS_COLUMNS = {
    "algorithmId": ColumnDefinition(
        name="algorithmId", display_name="Algorithm ID", data_type="string",
        required=True, unique=True,
        help_text="Unique identifier (ALG_NNN)",
        pattern=r"^ALG_\d{3}$",
        pattern_message="Algorithm ID should be in format ALG_NNN"
    ),
    "assessmentCode": ColumnDefinition(
        name="assessmentCode", display_name="Assessment Code", data_type="string",
        required=False,
        help_text="Assessment to trigger when condition is true"
    ),
    "moduleCode": ColumnDefinition(
        name="moduleCode", display_name="Module Code", data_type="string",
        required=False,
        help_text="Module context for this algorithm"
    ),
    "event": ColumnDefinition(
        name="event", display_name="Event Type", data_type="enum",
        required=True, default_value="ASSESSMENT",
        help_text="Type of action: ASSESSMENT, FINDING, MODULE"
    ),
    "priority": ColumnDefinition(
        name="priority", display_name="Priority", data_type="integer",
        required=False, default_value=1,
        help_text="Execution priority"
    ),
    "algorithm": ColumnDefinition(
        name="algorithm", display_name="Algorithm Expression", data_type="string",
        required=True,
        help_text="Logical expression using fact() and groups()"
    ),
    "findingCode": ColumnDefinition(
        name="findingCode", display_name="Finding Code", data_type="string",
        required=False,
        help_text="Finding to generate when condition is true"
    ),
    "changeControl": ColumnDefinition(
        name="changeControl", display_name="Change Control", data_type="string",
        required=False,
        help_text="Version control identifier"
    ),
}

ALGORITHMS_COLUMN_ORDER = [
    "algorithmId", "assessmentCode", "moduleCode", "event",
    "priority", "algorithm", "findingCode", "changeControl"
]

# ============================================================================
# ENUMERATIONS SHEET CONFIGURATION
# ============================================================================

ENUMERATIONS_COLUMNS = {
    "enumerationType": ColumnDefinition(
        name="enumerationType", display_name="Enumeration Type", data_type="string",
        required=True, max_length=64,
        help_text="Type identifier for the enumeration set (e.g., yesNo, frequency, severity)"
    ),
    "languageCode": ColumnDefinition(
        name="languageCode", display_name="Language Code", data_type="string",
        required=False, max_length=10,
        help_text="ISO language code for localization (e.g., en, es, fr). When empty, assumes default language."
    ),
    "seq": ColumnDefinition(
        name="seq", display_name="Sequence", data_type="integer",
        required=True,
        help_text="Display order sequence number. Lower numbers appear first."
    ),
    "value": ColumnDefinition(
        name="value", display_name="Value", data_type="string",
        required=True, max_length=255,
        help_text="The display text shown to the user (e.g., Yes, No, Not at all)"
    ),
    "derivedValue": ColumnDefinition(
        name="derivedValue", display_name="Derived Value", data_type="string",
        required=True, max_length=255,
        help_text="The computed value used in algorithm evaluation (e.g., true, false, numeric)"
    ),
    "tags": ColumnDefinition(
        name="tags", display_name="Tags", data_type="string",
        required=False,
        help_text="Additional tags (e.g., DELEGATE=factId)"
    ),
    "changeControl": ColumnDefinition(
        name="changeControl", display_name="Change Control", data_type="string",
        required=False,
        help_text="Audit trail and version tracking identifier"
    ),
}

ENUMERATIONS_COLUMN_ORDER = [
    "enumerationType", "languageCode", "seq", "value", "derivedValue", "tags", "changeControl"
]

# ============================================================================
# FINDINGS SHEET CONFIGURATION
# ============================================================================

FINDINGS_COLUMNS = {
    "findingCode": ColumnDefinition(
        name="findingCode", display_name="Finding Code", data_type="string",
        required=True, max_length=64, unique=True,
        help_text="Unique identifier for the clinical finding",
        pattern=r"^[A-Z0-9][A-Z0-9_\-]*[A-Z0-9]$|^[A-Z0-9]+$",
        pattern_message="Finding code should be UPPERCASE with underscores or hyphens"
    ),
    "name": ColumnDefinition(
        name="name", display_name="Name", data_type="string",
        required=True, max_length=255,
        help_text="Internal display name for the finding"
    ),
    "clientFriendlyName": ColumnDefinition(
        name="clientFriendlyName", display_name="Client Friendly Name", data_type="string",
        required=False,
        help_text="User-friendly name shown to clients"
    ),
    "icdCode": ColumnDefinition(
        name="icdCode", display_name="ICD Code", data_type="string",
        required=False, max_length=16,
        help_text="ICD-10 diagnostic code",
        pattern=r"^[A-Z]\d{2}(\.\d{1,4})?$",
        pattern_message="ICD code should be in format like F32.1 or A01.0"
    ),
    "tags": ColumnDefinition(
        name="tags", display_name="Tags", data_type="string",
        required=False,
        help_text="Comma-separated tags for categorization"
    ),
}

FINDINGS_COLUMN_ORDER = ["findingCode", "name", "clientFriendlyName", "icdCode", "tags"]

# ============================================================================
# FINDING RELATIONSHIPS SHEET CONFIGURATION
# ============================================================================

FINDING_RELATIONSHIPS_COLUMNS = {
    "sourceFindingCode": ColumnDefinition(
        name="sourceFindingCode", display_name="Source Finding Code", data_type="string",
        required=True, max_length=64,
        help_text="The source finding in the relationship"
    ),
    "targetFindingCode": ColumnDefinition(
        name="targetFindingCode", display_name="Target Finding Code", data_type="string",
        required=True, max_length=64,
        help_text="The target finding in the relationship"
    ),
    "sequence": ColumnDefinition(
        name="sequence", display_name="Sequence", data_type="integer",
        required=True,
        help_text="Order/position in the relationship chain. Lower numbers appear first."
    ),
    "relationshipTypeCode": ColumnDefinition(
        name="relationshipTypeCode", display_name="Relationship Type", data_type="string",
        required=True, max_length=32,
        help_text="Type of relationship (e.g., differential, comorbid, subtype, excludes, related)"
    ),
    "descriptor": ColumnDefinition(
        name="descriptor", display_name="Descriptor", data_type="string",
        required=False,
        help_text="Human-readable description of the relationship"
    ),
}

FINDING_RELATIONSHIPS_COLUMN_ORDER = [
    "sourceFindingCode", "targetFindingCode", "sequence", "relationshipTypeCode", "descriptor"
]

# ============================================================================
# SHEET REGISTRY
# ============================================================================

AVAILABLE_SHEETS = {
    "Modules": {
        "name": "Modules",
        "columns": MODULES_COLUMNS,
        "column_order": MODULES_COLUMN_ORDER,
        "description": "Define conversational modules used in the Insight CDS platform.",
        "icon": "📦"
    },
    "Assessments": {
        "name": "Assessments",
        "columns": ASSESSMENTS_COLUMNS,
        "column_order": ASSESSMENTS_COLUMN_ORDER,
        "description": "Define clinical assessments that collect information.",
        "icon": "📋"
    },
    "Level 0 Facts": {
        "name": "Level 0 Facts",
        "columns": LEVEL_FACTS_COLUMNS,
        "column_order": LEVEL_FACTS_COLUMN_ORDER,
        "description": LEVEL_FACTS_INFO[0]["description"],
        "icon": LEVEL_FACTS_INFO[0]["icon"],
        "level": 0
    },
    "Level 1 Facts": {
        "name": "Level 1 Facts",
        "columns": LEVEL_FACTS_COLUMNS,
        "column_order": LEVEL_FACTS_COLUMN_ORDER,
        "description": LEVEL_FACTS_INFO[1]["description"],
        "icon": LEVEL_FACTS_INFO[1]["icon"],
        "level": 1
    },
    "Level 2 Facts": {
        "name": "Level 2 Facts",
        "columns": LEVEL_FACTS_COLUMNS,
        "column_order": LEVEL_FACTS_COLUMN_ORDER,
        "description": LEVEL_FACTS_INFO[2]["description"],
        "icon": LEVEL_FACTS_INFO[2]["icon"],
        "level": 2
    },
    "Level 3 Facts": {
        "name": "Level 3 Facts",
        "columns": LEVEL_FACTS_COLUMNS,
        "column_order": LEVEL_FACTS_COLUMN_ORDER,
        "description": LEVEL_FACTS_INFO[3]["description"],
        "icon": LEVEL_FACTS_INFO[3]["icon"],
        "level": 3
    },
    "Algorithms": {
        "name": "Algorithms",
        "columns": ALGORITHMS_COLUMNS,
        "column_order": ALGORITHMS_COLUMN_ORDER,
        "description": "Define clinical decision rules that trigger assessments or findings.",
        "icon": "🔢"
    },
    "Enumerations": {
        "name": "Enumerations",
        "columns": ENUMERATIONS_COLUMNS,
        "column_order": ENUMERATIONS_COLUMN_ORDER,
        "description": "Define enumeration types and their values for ENUMERATION dataType fields.",
        "icon": "📝"
    },
    "Findings": {
        "name": "Findings",
        "columns": FINDINGS_COLUMNS,
        "column_order": FINDINGS_COLUMN_ORDER,
        "description": "Define clinical findings that can be generated by algorithms.",
        "icon": "🔍"
    },
    "Findings Relationships": {
        "name": "Findings Relationships",
        "columns": FINDING_RELATIONSHIPS_COLUMNS,
        "column_order": FINDING_RELATIONSHIPS_COLUMN_ORDER,
        "description": "Define relationships between clinical findings.",
        "icon": "🔗"
    },
}


def get_column_order(sheet_name: str) -> list:
    """Get column order for a sheet."""
    if sheet_name not in AVAILABLE_SHEETS:
        raise ValueError(f"Unknown sheet: {sheet_name}")
    return AVAILABLE_SHEETS[sheet_name]["column_order"]


def get_column_definitions(sheet_name: str) -> dict:
    """Get column definitions for a sheet."""
    if sheet_name not in AVAILABLE_SHEETS:
        raise ValueError(f"Unknown sheet: {sheet_name}")
    return AVAILABLE_SHEETS[sheet_name]["columns"]


def get_empty_row(sheet_name: str) -> dict:
    """Get an empty row with default values for a sheet."""
    if sheet_name not in AVAILABLE_SHEETS:
        raise ValueError(f"Unknown sheet: {sheet_name}")
    
    sheet_config = AVAILABLE_SHEETS[sheet_name]
    row = {}
    for col_name in sheet_config["column_order"]:
        col_def = sheet_config["columns"][col_name]
        row[col_name] = col_def.default_value
    return row


class ExcelManager:
    """Manages Excel file operations with support for multiple sheets."""
    
    def __init__(self):
        self.workbook: Optional[openpyxl.Workbook] = None
        self.file_path: Optional[Path] = None
        self.data: Dict[str, pd.DataFrame] = {}
    
    def load_from_file(self, file_source: BinaryIO) -> Dict[str, pd.DataFrame]:
        """Load an Excel file and read all recognized sheets."""
        self.workbook = openpyxl.load_workbook(file_source)
        self.data = {}
        
        for sheet_name in AVAILABLE_SHEETS.keys():
            if sheet_name in self.workbook.sheetnames:
                df = self._read_sheet(sheet_name)
                self.data[sheet_name] = df
            else:
                columns = get_column_order(sheet_name)
                self.data[sheet_name] = pd.DataFrame(columns=columns)
        
        return self.data
    
    def _read_sheet(self, sheet_name: str) -> pd.DataFrame:
        """Read a specific sheet from the workbook."""
        ws = self.workbook[sheet_name]
        headers = [cell.value for cell in ws[1]]
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                row_dict = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        value = row[i]
                        if isinstance(value, str) and value.upper() in ('TRUE', 'FALSE'):
                            value = value.upper() == 'TRUE'
                        row_dict[header] = value
                data.append(row_dict)
        
        expected_columns = get_column_order(sheet_name)
        df = pd.DataFrame(data)
        
        for col in expected_columns:
            if col not in df.columns:
                df[col] = None
        
        df = df[[col for col in expected_columns if col in df.columns]]
        return df
    
    def create_new_workbook(self) -> Dict[str, pd.DataFrame]:
        """Create a new empty workbook with all recognized sheets."""
        self.workbook = openpyxl.Workbook()
        self.data = {}
        
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        for sheet_name in AVAILABLE_SHEETS.keys():
            columns = get_column_order(sheet_name)
            self.data[sheet_name] = pd.DataFrame(columns=columns)
            ws = self.workbook.create_sheet(sheet_name)
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
        
        return self.data
    
    def update_sheet_data(self, sheet_name: str, df: pd.DataFrame):
        """Update the data for a specific sheet."""
        if sheet_name not in AVAILABLE_SHEETS:
            raise ValueError(f"Unknown sheet: {sheet_name}")
        self.data[sheet_name] = df.copy()
    
    def save_to_file(self, file_path: Optional[Path] = None) -> bytes:
        """Save the workbook to bytes."""
        if self.workbook is None:
            self.workbook = openpyxl.Workbook()
            if 'Sheet' in self.workbook.sheetnames:
                del self.workbook['Sheet']
        
        for sheet_name, df in self.data.items():
            self._write_sheet(sheet_name, df)
        
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.read()
    
    def _write_sheet(self, sheet_name: str, df: pd.DataFrame):
        """Write a DataFrame to a specific sheet with formatting."""
        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
        
        ws = self.workbook.create_sheet(sheet_name)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        columns = get_column_order(sheet_name)
        
        for col in columns:
            if col not in df.columns:
                df[col] = None
        df = df[columns]
        
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        ws.freeze_panes = "A2"

